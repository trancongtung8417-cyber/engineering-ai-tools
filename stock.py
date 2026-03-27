from spa_branding import setup_page
setup_page()   # ← phải đặt TRƯỚC mọi lệnh st.* khác

import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client, Client

# ─── CONFIG ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Quản Lý Kho Hàng",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── PWA + HIDE STREAMLIT UI ──────────────────────────────────────────────────
st.markdown("""
<head>
  <!-- PWA Meta Tags -->
  <meta name="mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Kho Hàng">
  <meta name="application-name" content="Kho Hàng">
  <meta name="theme-color" content="#1F4E79">

  <!-- PWA Icons (Apple & Android) -->
  <link rel="apple-touch-icon" href="https://i.postimg.cc/kgfrnncq/stock1.jpg">
  <link rel="icon" type="image/png" sizes="192x192" href="https://i.postimg.cc/kXsKsZs7/192-1.jpg">
  <link rel="icon" type="image/png" sizes="512x512" href="https://https://i.postimg.cc/tJynttVv/512.jpg">

  <!-- Web App Manifest (inline via JS) -->
  <script>
    const manifest = {
      name: "Quản Lý Kho Hàng",
      short_name: "Kho Hàng",
      description: "Hệ thống quản lý kho hàng chuyên nghiệp",
      start_url: "/",
      display: "standalone",
      background_color: "#0E1117",
      theme_color: "#1F4E79",
      orientation: "portrait-primary",
      icons: [
        {
          src: "https://img.icons8.com/fluency/192/warehouse.png",
          sizes: "192x192",
          type: "image/png",
          purpose: "any maskable"
        },
        {
          src: "https://img.icons8.com/fluency/512/warehouse.png",
          sizes: "512x512",
          type: "image/png",
          purpose: "any maskable"
        }
      ]
    };
    const blob = new Blob([JSON.stringify(manifest)], {type: "application/json"});
    const url  = URL.createObjectURL(blob);
    const link = document.createElement("link");
    link.rel   = "manifest";
    link.href  = url;
    document.head.appendChild(link);
  </script>
</head>

<style>
/* ── Ẩn toàn bộ Streamlit chrome ── */

/* Header bar (Fork · Star · GitHub · menu ⋮) */
header[data-testid="stHeader"],
[data-testid="stHeader"] { display: none !important; }

/* Nút Deploy (vương miện đỏ) */
[data-testid="stDeployButton"],
.stDeployButton { display: none !important; }

/* Toolbar góc dưới phải (logo Streamlit & status) */
[data-testid="stToolbar"],
.stToolbar,
#stDecoration,
footer,
footer > *,
[data-testid="stStatusWidget"],
.stStatusWidget { display: none !important; }

/* Xoá khoảng trắng thừa trên cùng do header để lại */
.main .block-container {
  padding-top: 1rem !important;
}

/* ── Style giao diện chuyên nghiệp ── */
[data-testid="stSidebar"] { background: #1F4E79; }
[data-testid="stSidebar"] * { color: #FFFFFF !important; }
.stButton>button {
  background: #1F4E79;
  color: white;
  border-radius: 8px;
  font-weight: 600;
}
.stButton>button:hover { background: #2980B9; }
h1 { color: #1F4E79; }
h2 { color: #2980B9; }
</style>

<!-- Banner "Thêm vào màn hình chính" cho iOS -->
<script>
  // Hiển thị hướng dẫn Add to Home Screen một lần duy nhất
  if (!localStorage.getItem("pwa_prompt_shown")) {
    const isIOS = /iphone|ipad|ipod/i.test(navigator.userAgent);
    const isStandalone = window.navigator.standalone === true
                      || window.matchMedia("(display-mode: standalone)").matches;
    if (isIOS && !isStandalone) {
      const div = document.createElement("div");
      div.id = "pwa-banner";
      div.style.cssText = `
        position:fixed; bottom:16px; left:50%; transform:translateX(-50%);
        background:#1F4E79; color:#fff; padding:12px 20px; border-radius:12px;
        font-size:14px; z-index:9999; box-shadow:0 4px 16px rgba(0,0,0,0.4);
        max-width:90vw; text-align:center; line-height:1.5;
      `;
      div.innerHTML = `
        📲 <strong>Thêm vào màn hình chính:</strong><br>
        Nhấn <strong>⬆ Share</strong> → <strong>Add to Home Screen</strong>
        <button onclick="document.getElementById('pwa-banner').remove();
                         localStorage.setItem('pwa_prompt_shown','1');"
                style="margin-left:12px;background:#fff;color:#1F4E79;
                       border:none;border-radius:6px;padding:4px 10px;
                       cursor:pointer;font-weight:700;">✕</button>
      `;
      document.body.appendChild(div);
      localStorage.setItem("pwa_prompt_shown", "1");
    }
  }

  // Android: lắng nghe sự kiện beforeinstallprompt
  let deferredPrompt;
  window.addEventListener("beforeinstallprompt", (e) => {
    e.preventDefault();
    deferredPrompt = e;
    const btn = document.createElement("button");
    btn.id = "android-install-btn";
    btn.innerText = "📲 Cài đặt App";
    btn.style.cssText = `
      position:fixed; bottom:20px; right:20px; z-index:9999;
      background:#1F4E79; color:#fff; border:none; border-radius:10px;
      padding:12px 18px; font-size:14px; font-weight:700; cursor:pointer;
      box-shadow:0 4px 12px rgba(0,0,0,0.3);
    `;
    btn.onclick = async () => {
      btn.remove();
      deferredPrompt.prompt();
      await deferredPrompt.userChoice;
      deferredPrompt = null;
    };
    document.body.appendChild(btn);
  });
</script>
""", unsafe_allow_html=True)

# ─── SUPABASE CONNECTION ────────────────────────────────────────────────────────
@st.cache_resource
def get_supabase() -> Client:
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

def sb() -> Client:
    return get_supabase()

# ─── DATA LAYER ────────────────────────────────────────────────────────────────

def fetch_warehouses():
    return sb().table("warehouses").select("*").order("id").execute().data or []

def fetch_products():
    return sb().table("products").select("*").order("id").execute().data or []

def fetch_transactions(wh_id=None, prod_id=None, txn_type=None, date_from=None, date_to=None):
    q = sb().table("transactions").select("*").order("date", desc=True).order("created_at", desc=True)
    if wh_id:    q = q.eq("warehouse_id", wh_id)
    if prod_id:  q = q.eq("product_id", prod_id)
    if txn_type: q = q.eq("type", txn_type)
    if date_from: q = q.gte("date", str(date_from))
    if date_to:   q = q.lte("date", str(date_to))
    return q.execute().data or []

def fetch_adjustments(wh_id=None, prod_id=None, date_from=None, date_to=None):
    q = sb().table("adjustments").select("*").order("date", desc=True).order("created_at", desc=True)
    if wh_id:    q = q.eq("warehouse_id", wh_id)
    if prod_id:  q = q.eq("product_id", prod_id)
    if date_from: q = q.gte("date", str(date_from))
    if date_to:   q = q.lte("date", str(date_to))
    return q.execute().data or []

def insert_warehouse(wid, name, location, desc):
    sb().table("warehouses").insert({"id": wid, "name": name, "location": location, "description": desc}).execute()

def delete_warehouse(wid):
    sb().table("warehouses").delete().eq("id", wid).execute()

def insert_product(pid, name, unit, category, desc):
    sb().table("products").insert({"id": pid, "name": name, "unit": unit, "category": category, "description": desc}).execute()

def delete_product(pid):
    sb().table("products").delete().eq("id", pid).execute()

def insert_transaction(txn_date, wh_id, prod_id, txn_type, qty, note):
    sb().table("transactions").insert({
        "date": str(txn_date), "warehouse_id": wh_id, "product_id": prod_id,
        "type": txn_type, "quantity": qty, "note": note,
        "created_at": datetime.now().isoformat(),
    }).execute()

def insert_adjustment(adj_date, wh_id, prod_id, delta, old_stock, new_stock, reason):
    sb().table("adjustments").insert({
        "date": str(adj_date), "warehouse_id": wh_id, "product_id": prod_id,
        "delta": delta, "old_stock": old_stock, "new_stock": new_stock,
        "reason": reason, "created_at": datetime.now().isoformat(),
    }).execute()

# ─── BUSINESS LOGIC ────────────────────────────────────────────────────────────

def get_stock(wh_id, prod_id, up_to_date=None):
    q_t = sb().table("transactions").select("type,quantity").eq("warehouse_id", wh_id).eq("product_id", prod_id)
    q_a = sb().table("adjustments").select("delta").eq("warehouse_id", wh_id).eq("product_id", prod_id)
    if up_to_date:
        q_t = q_t.lte("date", str(up_to_date))
        q_a = q_a.lte("date", str(up_to_date))
    txns = q_t.execute().data or []
    adjs = q_a.execute().data or []
    balance = sum(t["quantity"] if t["type"] == "Nhập" else -t["quantity"] for t in txns)
    balance += sum(a["delta"] for a in adjs)
    return balance

def wh_name(wid, warehouses):
    return next((w["name"] for w in warehouses if w["id"] == wid), wid)

def prod_name(pid, products):
    return next((p["name"] for p in products if p["id"] == pid), pid)

# ─── EXCEL EXPORT ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
IMPORT_FILL = PatternFill("solid", start_color="D6E4F0")
EXPORT_FILL = PatternFill("solid", start_color="FADBD8")
ADJ_FILL    = PatternFill("solid", start_color="D5F5E3")

def tb():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def sh(cell):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = tb()

def sc(cell, fill=None, align="center"):
    cell.font = Font(name="Arial", size=10)
    if fill: cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = tb()

def auto_width(ws):
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

def build_excel(date_from, date_to):
    warehouses = fetch_warehouses()
    products   = fetch_products()
    txns_all   = fetch_transactions(date_from=date_from, date_to=date_to)
    adjs_all   = fetch_adjustments(date_from=date_from, date_to=date_to)
    prev_date  = (pd.to_datetime(str(date_from)) - timedelta(days=1)).date()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Tổng hợp
    ws1 = wb.create_sheet("Tổng Hợp Tồn Kho")
    h1 = ["Kho","Mã Hàng","Tên Hàng","Tồn Đầu Kỳ","Tổng Nhập","Tổng Xuất","Điều Chỉnh","Tồn Cuối Kỳ"]
    ws1.append(h1); ws1.row_dimensions[1].height = 30
    for i in range(1, len(h1)+1): sh(ws1.cell(1, i))
    r = 2
    for w in warehouses:
        for p in products:
            wid, pid = w["id"], p["id"]
            open_s = get_stock(wid, pid, str(prev_date))
            imp = sum(t["quantity"] for t in txns_all if t["warehouse_id"]==wid and t["product_id"]==pid and t["type"]=="Nhập")
            exp = sum(t["quantity"] for t in txns_all if t["warehouse_id"]==wid and t["product_id"]==pid and t["type"]=="Xuất")
            adj = sum(a["delta"] for a in adjs_all if a["warehouse_id"]==wid and a["product_id"]==pid)
            vals = [w["name"], p["id"], p["name"], open_s, imp, exp, adj, open_s+imp-exp+adj]
            ws1.append(vals)
            for c in range(1, 9): sc(ws1.cell(r, c), align="right" if c>3 else "left")
            r += 1
    auto_width(ws1); ws1.freeze_panes = "A2"

    # Sheet 2: Chi tiết giao dịch
    ws2 = wb.create_sheet("Chi Tiết Giao Dịch")
    h2 = ["Ngày","Kho","Loại","Mã Hàng","Tên Hàng","Số Lượng","Đơn Vị","Ghi Chú"]
    ws2.append(h2); ws2.row_dimensions[1].height = 30
    for i in range(1, len(h2)+1): sh(ws2.cell(1, i))
    r = 2
    for t in sorted(txns_all, key=lambda x: x["date"]):
        fill = IMPORT_FILL if t["type"]=="Nhập" else EXPORT_FILL
        unit = next((p["unit"] for p in products if p["id"]==t["product_id"]), "")
        vals = [t["date"], wh_name(t["warehouse_id"], warehouses), t["type"],
                t["product_id"], prod_name(t["product_id"], products), t["quantity"], unit, t.get("note","")]
        ws2.append(vals)
        for c in range(1, 9): sc(ws2.cell(r, c), fill=fill, align="right" if c==6 else "left")
        r += 1
    auto_width(ws2); ws2.freeze_panes = "A2"

    # Sheet 3: Điều chỉnh
    ws3 = wb.create_sheet("Điều Chỉnh Tồn Kho")
    h3 = ["Ngày","Kho","Mã Hàng","Tên Hàng","Điều Chỉnh (+/-)","Lý Do"]
    ws3.append(h3); ws3.row_dimensions[1].height = 30
    for i in range(1, len(h3)+1): sh(ws3.cell(1, i))
    r = 2
    for a in sorted(adjs_all, key=lambda x: x["date"]):
        vals = [a["date"], wh_name(a["warehouse_id"], warehouses),
                a["product_id"], prod_name(a["product_id"], products), a["delta"], a.get("reason","")]
        ws3.append(vals)
        for c in range(1, 7): sc(ws3.cell(r, c), fill=ADJ_FILL, align="right" if c==5 else "left")
        r += 1
    auto_width(ws3)

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ─── SIDEBAR ───────────────────────────────────────────────────────────────────
st.sidebar.markdown("## 🏭 Quản Lý Kho Hàng")
menu = st.sidebar.radio("", [
    "📊 Tổng Quan",
    "🏪 Quản Lý Kho",
    "📦 Quản Lý Hàng Hóa",
    "📥 Nhập / Xuất Hàng",
    "🔧 Điều Chỉnh Tồn",
    "📋 Lịch Sử Giao Dịch",
    "📤 Xuất Excel",
])

# ══════════════════════════════════════════════════════════════════════════════
if menu == "📊 Tổng Quan":
    st.title("📊 Tổng Quan Kho Hàng")
    warehouses   = fetch_warehouses()
    products     = fetch_products()
    transactions = fetch_transactions()
    adjustments  = fetch_adjustments()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("🏪 Số Kho", len(warehouses))
    c2.metric("📦 Mã Hàng", len(products))
    c3.metric("📋 Giao Dịch", len(transactions))
    c4.metric("🔧 Điều Chỉnh", len(adjustments))

    if warehouses and products:
        st.subheader("📦 Tồn Kho Hiện Tại")
        with st.spinner("Đang tải..."):
            rows = []
            for w in warehouses:
                for p in products:
                    stock = get_stock(w["id"], p["id"])
                    rows.append({
                        "Kho": w["name"], "Mã Hàng": p["id"], "Tên Hàng": p["name"],
                        "Đơn Vị": p["unit"], "Tồn Kho": stock,
                        "Trạng Thái": "✅ Bình thường" if stock > 0 else ("⚠️ Hết hàng" if stock == 0 else "🔴 Âm kho"),
                    })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        st.subheader("📈 Giao Dịch 30 Ngày Gần Nhất")
        recent = fetch_transactions(date_from=date.today() - timedelta(days=30))
        if recent:
            r1, r2 = st.columns(2)
            r1.metric("📥 Tổng Nhập", f"{sum(t['quantity'] for t in recent if t['type']=='Nhập'):,}")
            r2.metric("📤 Tổng Xuất", f"{sum(t['quantity'] for t in recent if t['type']=='Xuất'):,}")
        else:
            st.info("Chưa có giao dịch trong 30 ngày gần nhất.")
    else:
        st.info("👋 Hãy bắt đầu bằng cách **thêm kho** và **thêm hàng hóa** ở menu bên trái!")

elif menu == "🏪 Quản Lý Kho":
    st.title("🏪 Quản Lý Kho")
    warehouses = fetch_warehouses()

    with st.expander("➕ Thêm Kho Mới", expanded=True):
        with st.form("add_wh"):
            c1, c2 = st.columns(2)
            wid   = c1.text_input("Mã Kho *", placeholder="KHO01")
            wname = c2.text_input("Tên Kho *", placeholder="Kho Hà Nội")
            wloc  = st.text_input("Địa Chỉ")
            wdesc = st.text_area("Ghi Chú", height=60)
            if st.form_submit_button("💾 Lưu Kho"):
                if not wid or not wname:
                    st.error("Vui lòng nhập Mã Kho và Tên Kho!")
                elif any(w["id"] == wid for w in warehouses):
                    st.error(f"Mã kho '{wid}' đã tồn tại!")
                else:
                    insert_warehouse(wid, wname, wloc, wdesc)
                    st.success(f"✅ Đã thêm kho '{wname}'!"); st.rerun()

    st.subheader("📋 Danh Sách Kho")
    if warehouses:
        df = pd.DataFrame(warehouses)[["id","name","location","description"]]
        df.columns = ["Mã Kho","Tên Kho","Địa Chỉ","Ghi Chú"]
        st.dataframe(df, use_container_width=True, hide_index=True)
        wh_del = st.selectbox("Chọn kho cần xóa", [f"{w['id']} - {w['name']}" for w in warehouses])
        if st.button("🗑️ Xóa Kho"):
            delete_warehouse(wh_del.split(" - ")[0])
            st.success("Đã xóa!"); st.rerun()
    else:
        st.info("Chưa có kho nào.")

elif menu == "📦 Quản Lý Hàng Hóa":
    st.title("📦 Quản Lý Hàng Hóa")
    products = fetch_products()

    with st.expander("➕ Thêm Mã Hàng Mới", expanded=True):
        with st.form("add_prod"):
            c1, c2, c3 = st.columns(3)
            pid   = c1.text_input("Mã Hàng *", placeholder="SP001")
            pname = c2.text_input("Tên Hàng *", placeholder="iPhone 15")
            punit = c3.text_input("Đơn Vị *", placeholder="Cái")
            c4, c5 = st.columns(2)
            pcat  = c4.text_input("Nhóm Hàng")
            pdesc = c5.text_input("Ghi Chú")
            if st.form_submit_button("💾 Lưu Hàng"):
                if not pid or not pname or not punit:
                    st.error("Nhập đầy đủ Mã, Tên, Đơn vị!")
                elif any(p["id"] == pid for p in products):
                    st.error(f"Mã '{pid}' đã tồn tại!")
                else:
                    insert_product(pid, pname, punit, pcat, pdesc)
                    st.success(f"✅ Đã thêm '{pname}'!"); st.rerun()

    st.subheader("📋 Danh Sách Hàng Hóa")
    if products:
        df = pd.DataFrame(products)[["id","name","unit","category","description"]]
        df.columns = ["Mã Hàng","Tên Hàng","Đơn Vị","Nhóm","Ghi Chú"]
        st.dataframe(df, use_container_width=True, hide_index=True)
        pd_del = st.selectbox("Chọn hàng cần xóa", [f"{p['id']} - {p['name']}" for p in products])
        if st.button("🗑️ Xóa Hàng"):
            delete_product(pd_del.split(" - ")[0])
            st.success("Đã xóa!"); st.rerun()
    else:
        st.info("Chưa có mã hàng nào.")

elif menu == "📥 Nhập / Xuất Hàng":
    st.title("📥 Nhập / Xuất Hàng")
    warehouses = fetch_warehouses()
    products   = fetch_products()
    if not warehouses: st.warning("⚠️ Thêm kho trước!"); st.stop()
    if not products:   st.warning("⚠️ Thêm hàng hóa trước!"); st.stop()

    for tab, ttype in zip(st.tabs(["📥 Nhập Hàng", "📤 Xuất Hàng"]), ["Nhập", "Xuất"]):
        with tab:
            with st.form(f"form_{ttype}"):
                c1, c2 = st.columns(2)
                txn_date = c1.date_input("Ngày *", value=date.today())
                wh_sel   = c2.selectbox("Kho *", [f"{w['id']} - {w['name']}" for w in warehouses])
                c3, c4 = st.columns(2)
                pr_sel = c3.selectbox("Hàng hóa *", [f"{p['id']} - {p['name']}" for p in products])
                qty    = c4.number_input("Số Lượng *", min_value=1, value=1)
                note   = st.text_input("Ghi Chú / Chứng Từ")
                if st.form_submit_button(f"✅ Xác Nhận {ttype} Hàng"):
                    wid = wh_sel.split(" - ")[0]
                    pid = pr_sel.split(" - ")[0]
                    cur = get_stock(wid, pid)
                    if ttype == "Xuất" and qty > cur:
                        st.error(f"❌ Không đủ hàng! Tồn: {cur:,}")
                    else:
                        insert_transaction(txn_date, wid, pid, ttype, qty, note)
                        st.success(f"✅ {ttype} {qty:,} thành công! Tồn mới: {get_stock(wid, pid):,}")
                        st.rerun()

elif menu == "🔧 Điều Chỉnh Tồn":
    st.title("🔧 Điều Chỉnh Tồn Kho")
    st.info("Dùng khi kiểm kê phát hiện chênh lệch, hàng hỏng, mất mát...")
    warehouses = fetch_warehouses()
    products   = fetch_products()
    if not warehouses or not products: st.warning("Cần có kho và hàng trước!"); st.stop()

    with st.form("adj_form"):
        c1, c2 = st.columns(2)
        adj_date = c1.date_input("Ngày Điều Chỉnh", value=date.today())
        wh_sel   = c2.selectbox("Kho", [f"{w['id']} - {w['name']}" for w in warehouses])
        pr_sel   = st.selectbox("Hàng Hóa", [f"{p['id']} - {p['name']}" for p in products])
        wid = wh_sel.split(" - ")[0]
        pid = pr_sel.split(" - ")[0]
        cur = get_stock(wid, pid)
        st.markdown(f"**Tồn kho hiện tại:** `{cur:,}`")
        c3, c4 = st.columns(2)
        new_qty = c3.number_input("Số Lượng Thực Tế", min_value=0, value=max(0, cur))
        reason  = c4.text_input("Lý Do *", placeholder="Kiểm kê, hàng hỏng...")
        delta = new_qty - cur
        if delta > 0:   st.success(f"Cộng thêm +{delta:,}")
        elif delta < 0: st.warning(f"Trừ đi {delta:,}")
        else:           st.info("Không thay đổi")
        if st.form_submit_button("💾 Lưu Điều Chỉnh"):
            if not reason: st.error("Nhập lý do điều chỉnh!")
            elif delta == 0: st.info("Không có thay đổi.")
            else:
                insert_adjustment(adj_date, wid, pid, delta, cur, new_qty, reason)
                st.success(f"✅ Tồn mới: {new_qty:,}"); st.rerun()

elif menu == "📋 Lịch Sử Giao Dịch":
    st.title("📋 Lịch Sử Giao Dịch")
    warehouses = fetch_warehouses()
    products   = fetch_products()
    tab1, tab2 = st.tabs(["📋 Nhập / Xuất", "🔧 Điều Chỉnh"])

    with tab1:
        f1, f2, f3 = st.columns(3)
        sel_wh = f1.selectbox("Kho",  ["Tất cả"] + [f"{w['id']} - {w['name']}" for w in warehouses], key="fwh")
        sel_pr = f2.selectbox("Hàng", ["Tất cả"] + [f"{p['id']} - {p['name']}" for p in products], key="fpr")
        sel_tp = f3.selectbox("Loại", ["Tất cả","Nhập","Xuất"], key="ftp")
        txns = fetch_transactions(
            wh_id    = sel_wh.split(" - ")[0] if sel_wh != "Tất cả" else None,
            prod_id  = sel_pr.split(" - ")[0] if sel_pr != "Tất cả" else None,
            txn_type = sel_tp if sel_tp != "Tất cả" else None,
        )
        if txns:
            rows = [{"Ngày":t["date"],"Kho":wh_name(t["warehouse_id"],warehouses),"Loại":t["type"],
                     "Mã Hàng":t["product_id"],"Tên Hàng":prod_name(t["product_id"],products),
                     "Số Lượng":t["quantity"],"Ghi Chú":t.get("note","")} for t in txns]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            st.caption(f"Tổng: {len(rows)} giao dịch")
        else: st.info("Không có giao dịch phù hợp.")

    with tab2:
        adjs = fetch_adjustments(
            wh_id   = sel_wh.split(" - ")[0] if sel_wh != "Tất cả" else None,
            prod_id = sel_pr.split(" - ")[0] if sel_pr != "Tất cả" else None,
        )
        if adjs:
            rows = [{"Ngày":a["date"],"Kho":wh_name(a["warehouse_id"],warehouses),
                     "Mã Hàng":a["product_id"],"Tên Hàng":prod_name(a["product_id"],products),
                     "Tồn Cũ":a["old_stock"],"Tồn Mới":a["new_stock"],
                     "Chênh Lệch":a["delta"],"Lý Do":a.get("reason","")} for a in adjs]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else: st.info("Chưa có điều chỉnh nào.")

elif menu == "📤 Xuất Excel":
    st.title("📤 Xuất Báo Cáo Excel")
    c1, c2 = st.columns(2)
    d_from = c1.date_input("Từ Ngày", value=date(date.today().year, 1, 1))
    d_to   = c2.date_input("Đến Ngày", value=date.today())
    if d_from > d_to:
        st.error("Ngày bắt đầu phải nhỏ hơn ngày kết thúc!")
    else:
        if st.button("📥 Tạo File Excel", use_container_width=True):
            with st.spinner("Đang tạo từ Supabase..."):
                buf = build_excel(d_from, d_to)
            st.download_button(
                "⬇️ Tải Xuống Excel", data=buf,
                file_name=f"KhoHang_{d_from}_{d_to}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.success("✅ File Excel đã sẵn sàng!")
